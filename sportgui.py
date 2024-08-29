
from PyQt5.QtWidgets import *
from PyQt5.QtMultimedia import *
from PyQt5.QtWidgets import QApplication
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from Ui_sport import Ui_MainWindow
# from myVideoWidget import myVideoWidget
import sys
from PyQt5 import QtWidgets
import time
from PyQt5.QtCore import QDate, QDateTime
# from services.sport.sportquery
import datetime
from dao import sportinfo
from PyQt5.QtWidgets import QApplication, QMainWindow, QTimeEdit
from openpyxl.styles import Font, colors
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import colors
import openpyxl
from openpyxl.styles import Border, Side

class sportMainWindow(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(Ui_MainWindow, self).__init__()
        self.setupUi(self)
        self.querydata.clicked.connect(self.querysportinfo)   # 打开视频文件按钮
        self.exportdata.clicked.connect(self.savedata)   # 打开视频文件按钮
        self.pathselect.clicked.connect(self.selfilepath)   # 打开视频文件按钮
        
       # self.lineTimeInfo.clicked.connect(self.show_time_dialog)   # 打开视频文件按钮
        # self.lineTimeInfo.setPlaceholderText("Click Please")

        # self.lineTimeInfo.mousePressEvent = self.show_time_dialog
        self.fromDateTime.setCalendarPopup(True)
        # self.dateTimeEdit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.fromDateTime.setDisplayFormat("yyyy-MM-dd HH:mm")
        self.fromDateTime.setDateTime(datetime.datetime.now() + datetime.timedelta(hours=-24))

        self.toDateTime.setCalendarPopup(True)
        self.toDateTime.setDisplayFormat("yyyy-MM-dd HH:mm")
        self.toDateTime.setDateTime(QDateTime.currentDateTime())
        self.destpath = "D:/sample.xlsx"
        self.savepath.setText(self.destpath)
       
        # # 设置最小日期
        # self.dateTimeEdit.setMinimumDate(QDate.currentDate().addDays(-365))
        # # 设置最大日期
        # self.dateTimeEdit.setMaximumDate(QDate.currentDate().addDays(365))
        # # 允许弹出日历控件
        
          # 设置时间部分可见
        # self.dateTimeEdit.setTimeSpec(Qt.TimeSpec.TimeSpec_Second)

        # self.dateTimeEdit.dateChanged.connect(self.onDateChanged)
        # self.dateTimeEdit.dateTimeChanged.connect(self.onDateTimeChanged)
        # self.dateTimeEdit.timeChanged.connect(self.onTimeChanged)
        # for i in range(12):
        #     self.tableWidget.insertRow(i)
        #     for j in range(12):
                
        #         item = QtWidgets.QTableWidgetItem("111222fffbbb"+str(i))
        #         self.tableWidget.setItem(i, j, item)
        # self.sld_video_pressed=False  #判断当前进度条识别否被鼠标点击
            
    def show_time_dialog(self, mouseEvent):
        time_edit, _ = QTimeEdit.getTime(self, "选择时间", Qt.currentDateTime())
        self.lineTimeInfo.setPlaceholderText("Thanks Man :)")
        print(time_edit.time())

     # 日期发生改变时执行
    def onDateChanged(self, date):
        print(date)
 
    # 无论日期还是时间发生改变，都会执行
    def onDateTimeChanged(self, dateTime):
        print(dateTime)
 
    # 时间发生改变时执行
    def onTimeChanged(self, time):
        print(time)

    def selfilepath(self):
        self.destpath, filetype = QFileDialog.getSaveFileName(self, "文件保存", "myfile.xlsx", "Excel工作簿 (*.xlsx)")
        if self.destpath:  # 如果获取的路径非空
            print("destpath=", self.destpath)
            self.savepath.setText(self.destpath)

    def setsheetwidth(self, ws):
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 7
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 22
        ws.column_dimensions['O'].width = 24
        ws.column_dimensions['P'].width = 24

    def savedata(self):
        filepath = self.savepath.text()
        print("exportfile: func begin text=", filepath)
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active # 获取当前活动的工作表，默认情况下，活动工作表是Excel文件中的第一个工作表，也可以通 
        ws.title = "sheet";
        ws["A1"] = "序列"
        ws["B1"] = "年月日"
        ws["C1"] = "星期"
        ws["D1"] = "比赛时间"
        ws["E1"] = "联赛名"
        ws["F1"] = "主队名"
        ws["G1"] = "客队名"
        ws["H1"] = "主盘口"
        ws["I1"] = "主赔率"
        ws["J1"] = "客盘口"
        ws["K1"] = "客赔率"
        ws["L1"] = "全场比分"
        ws["M1"] = "指数运算结果"
        ws["N1"] = "新闻首发时间"
        ws["O1"] = "新闻主标题"
        ws["P1"] = "新闻内容"
        self.setsheetwidth(ws)
        from_time_str = self.fromDateTime.dateTime().toString('yyyy-MM-dd hh:mm')
        to_time_str = self.toDateTime.dateTime().toString('yyyy-MM-dd hh:mm')
        print("date_time_str=", from_time_str+":00", " to_time_str=", to_time_str+":00")
        leaguename = self.leaguename.text()
        print("leaguename=", leaguename)
        newtitle = self.newtitle.text()
        arrleaguename = []
        if leaguename != "":
            arrtmp =  leaguename.split(":")
            for data in arrtmp:
                print("data=", data)
                arrleaguename.append(data)
        
        # print("size=", len(arrleaguename), " arrtmplen=", len(arrtmp))
        arroddsdata = sportinfo.query_odds_data(from_time_str+":00", to_time_str+":00", arrleaguename, newtitle)
        # week_list = ["星期一","星期二","星期三","星期四","星期五","星期六","星期日"]
        week_list = ["一","二","三","四","五","六","日"]
        for i in range(len(arroddsdata)):
            stridx = str(i + 2)
            ws["A"+stridx] = i+1
            start_time = str(arroddsdata[i]['start_time'])
            arrstarttime = start_time.split(" ")
            print("0=", arrstarttime[0], " 1=", arrstarttime[1])
            if len(arroddsdata) > 0:
                ws["B"+stridx] = arrstarttime[0].replace('-', '/')
            if len(arroddsdata) > 1:
                ws["D"+stridx] = arrstarttime[1]
            arrdate = arrstarttime[0].split("-")
            if len(arrdate) > 2:
                date_obj = datetime.date(int(arrdate[0]), int(arrdate[1]), int(arrdate[2]))
                # 获取星期几的信息
                weekday = date_obj.weekday()
                print("weekday=", weekday+1)
                ws["C"+stridx] = week_list[weekday]

            ws["E"+stridx] = arroddsdata[i]['eventname']
            ws["F"+stridx] = arroddsdata[i]['HomeName']
            ws["G"+stridx] = arroddsdata[i]['AwayName']
            arrtmp = arroddsdata[i]['HandicapHomehcapdisp'].split("/")
            if len(arrtmp) > 1:
                strres = str(-float(arrtmp[0])) + "/" + str(-float(arrtmp[1]))
                ws["H"+stridx] = strres
            else:
                ws["H"+stridx] = -float(arroddsdata[i]['HandicapHomehcapdisp'])
            ws["I"+stridx] = arroddsdata[i]['HandicapHomeRate']
            ws["J"+stridx] = arroddsdata[i]['Handicapawayhcapdisp']
            ws["K"+stridx] = arroddsdata[i]['HandicapawayRate']
            ws["L"+stridx] = arroddsdata[i]['currentscore']
            ws["M"+stridx] = ""
            ws["N"+stridx] = str(arroddsdata[i]['reportTime'])
            ws["O"+stridx] = arroddsdata[i]['newstitle']
            ws["P"+stridx] = arroddsdata[i]['newscontent']
        wb.save(filepath)
        QMessageBox.information(self, '提示信息', '导出数据到文件成功')  # 可以不带选择按键，也可以带ok按键
    def querysportinfo(self):        
        from_time_str = self.fromDateTime.dateTime().toString('yyyy-MM-dd hh:mm')
        to_time_str = self.toDateTime.dateTime().toString('yyyy-MM-dd hh:mm')
        print("date_time_str=", from_time_str+":00", " to_time_str=", to_time_str+":00")
        leaguename = self.leaguename.text()
        print("leaguename=", leaguename)
        newtitle = self.newtitle.text()
        arrleaguename = []
        if leaguename != "":
            arrtmp =  leaguename.split(":")
            for data in arrtmp:
                print("data=", data)
                arrleaguename.append(data)
        
        # print("size=", len(arrleaguename), " arrtmplen=", len(arrtmp))
        arroddsdata = sportinfo.query_odds_data(from_time_str+":00", to_time_str+":00", arrleaguename, newtitle)
        
        print("leaguename=", leaguename)
        self.tableWidget.setRowCount(0);
        for i in range(len(arroddsdata)):
            self.tableWidget.insertRow(i)
            item = QtWidgets.QTableWidgetItem(str(arroddsdata[i]['start_time']))
            self.tableWidget.setItem(i, 0, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['eventname'])
            self.tableWidget.setItem(i, 1, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['HomeName'])
            self.tableWidget.setItem(i, 2, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['AwayName'])
            self.tableWidget.setItem(i, 3, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['HandicapHomehcapdisp'])
            self.tableWidget.setItem(i, 4, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['HandicapHomeRate'])
            self.tableWidget.setItem(i, 5, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['Handicapawayhcapdisp'])
            self.tableWidget.setItem(i, 6, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['HandicapawayRate'])
            self.tableWidget.setItem(i, 7, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['currentscore'])
            self.tableWidget.setItem(i, 8, item)

            item = QtWidgets.QTableWidgetItem(str(arroddsdata[i]['reportTime']))
            self.tableWidget.setItem(i, 10, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['newstitle'])
            self.tableWidget.setItem(i, 11, item)

            item = QtWidgets.QTableWidgetItem(arroddsdata[i]['newscontent'])
            self.tableWidget.setItem(i, 12, item)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    vieo_gui = sportMainWindow()
    vieo_gui.show()
    sys.exit(app.exec_())


